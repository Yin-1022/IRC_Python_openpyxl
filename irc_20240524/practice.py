from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.create_sheet("name")
ws = wb["name"]
# print(ws)
# print(wb.sheetnames)
wb.save("test.xlsx")

ws = wb.active
ws.title = "my sheet"
ws["A1"].value = 123
ws["A2"].value = 456
ws["B1"].value = "abc"
ws["d2"] = "Eason"
ws.append([11, 22, 33, 44, 55])
wb.save("test.xlsx")

# print(ws)
# print(ws["A2"])
# print(ws["A2"].value)

wb = load_workbook("test.xlsx")
ws = wb.active

for row in range(1, ws.max_row+1):
    for col in range(1, ws.max_column+1):
        char = get_column_letter(col)

        if(ws[char+str(row)].value != None):
            print(f"{char}{row}={ws[char+str(row)].value}")

#merge.cells("range1 : range2")
#ws.insert/delete_cols()/rows()



wb = load_workbook("test.xlsx")
ws = wb.active
# ws.move_range("C2:D3", rows=4, cols=-1)
# ws.insert_cols(2)
wb.save("test.xlsx")